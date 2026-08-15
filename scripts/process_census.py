import os
import re
import openpyxl
import pandas as pd

os.chdir(os.path.dirname(os.path.abspath(__file__)) + "/..")

SRC = {
    "ahmedabad": "raw/census/ahmedabad_PCA_ward.xlsx",
    "gandhinagar": "raw/census/gandhinagar_PCA_ward.xlsx",
}
TARGET_TOWN = {
    "ahmedabad": "Ahmadabad (M Corp.)",
    "gandhinagar": "Gandhinagar (NA)",
}
OUT = {
    "ahmedabad": "refined/census_ward_population_ahmedabad.csv",
    "gandhinagar": "refined/census_ward_population_gandhinagar.csv",
}
DISTRICT_POP = {  # census 2011 district totals (from same files, DISTRICT level)
    "ahmedabad": 7214225,
    "gandhinagar": 1391753,
}
# Compound growth rate applied to project 2026 population (approx. AHM/GNR growth trend)
CAGR = {  # ~2.2%/yr AHM urban, ~2.5%/yr GNR urban
    "ahmedabad": 0.022,
    "gandhinagar": 0.025,
}
YEARS = 2026 - 2011

TOWN_NAME = re.compile(r"^(.*?)\s*WARD NO\.-\d+.*$")

FIELDS = [
    "No_HH", "TOT_P", "TOT_M", "TOT_F", "P_06", "M_06", "F_06",
    "P_SC", "M_SC", "F_SC", "P_ST", "M_ST", "F_ST",
    "P_LIT", "M_LIT", "F_LIT", "P_ILL", "M_ILL", "F_ILL",
    "TOT_WORK_P", "TOT_WORK_M", "TOT_WORK_F",
    "MAINWORK_P", "MAINWORK_M", "MAINWORK_F",
    "MAIN_CL_P", "MAIN_AL_P", "MAIN_HH_P", "MAIN_OT_P",
    "MARGWORK_P", "MARGWORK_M", "MARGWORK_F",
    "NON_WORK_P", "NON_WORK_M", "NON_WORK_F",
]


def load_wards(path, target_town):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for r in rows:
        if str(r[idx["Level"]] or "") != "WARD":
            continue
        town = str(r[idx["Name"]] or "")
        if not town.startswith(target_town):
            continue
        ward = int(str(r[idx["Ward"]] or 0).replace("000", ""))
        rec = {"town": town, "ward": ward}
        for f in FIELDS:
            rec[f] = int(r[idx[f]] or 0)
        out.append(rec)
    wb.close()
    out.sort(key=lambda x: x["ward"])
    return out


def main():
    os.makedirs("refined", exist_ok=True)
    for city in ("ahmedabad", "gandhinagar"):
        wards = load_wards(SRC[city], TARGET_TOWN[city])
        df = pd.DataFrame(wards)
        df["city"] = city.capitalize()
        df = df[["city", "town", "ward"] + FIELDS]
        # projections (est, not official)
        df["pop_2026_est"] = (df["TOT_P"] * (1 + CAGR[city]) ** YEARS).round(0).astype(int)
        df.to_csv(OUT[city], index=False)
        tot = int(df["TOT_P"].sum())
        print(f"{city}: {len(df)} wards, sum pop = {tot:,}, file {OUT[city]}")
        print("  ward rows sample:")
        print(df[["ward", "TOT_P", "No_HH", "P_LIT", "pop_2026_est"]].head(3).to_string(index=False))


if __name__ == "__main__":
    main()
